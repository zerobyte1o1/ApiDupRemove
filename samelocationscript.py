
from openpyxl import load_workbook,Workbook
import os,time
class Teletraan:
    def __init__(self):
        self.urls=[]
        if not os.path.exists('interface.xlsx'):
            self.wb = Workbook()
        else:
            self.wb = load_workbook("interface.xlsx")


        print("请输入公有路由url")
        self.url=input()
        print("请输入模块/项目名")
        sheetname=input()
        self.ws = self.wb.create_sheet(sheetname, 0)
        self.ws.append(["url"])
    def request(self,flow):
        # url="https://mc-test.teletraan.io/graphql/?"
        if self.url in flow.request.url:
            if flow.request.url not in self.urls:
                self.urls.append(flow.request.url)
                # print(self.urls)

                self.ws.append([flow.request.url])
                self.wb.save("interface.xlsx")
                print(flow.request.url)

addons=[
    Teletraan()
]

if __name__ == '__main__':
    os.system("mitmdump -q -p 8899 -s samelocationscript.py")